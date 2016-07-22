#m_controller
import wx
import wx.grid
import wx.aui
import wx.lib.agw.aui
import os
import view
import model
import threading
import sys
import time
from wx import SpinButton
from model import Model
from view import View
import openpyxl
import openpyxl.styles as xl_sty
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from collections import OrderedDict
import threading as Thd
import time
import matplotlib as mpl
mpl.use('WXAgg')
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.backends.backend_wxagg import FigureCanvasWxAgg as Canvas
from matplotlib.backends.backend_wxagg import NavigationToolbar2Wx as Toolbar

class PlotDataPage(wx.Panel):
    def __init__(self, parent, page_name):
        self.mParent = parent
        self.mPage = wx.Panel(parent, 
                         wx.ID_ANY, 
                         wx.DefaultPosition, 
                         wx.DefaultSize, 
                         wx.TAB_TRAVERSAL)
        pageSizer = wx.BoxSizer(wx.VERTICAL)
        
        
        self.mFig = plt.figure()
        
        
        
        
                
        self.mCanvas = Canvas(self.mPage, -1, self.mFig)
        self.mCanvas.mpl_connect('motion_notify_event', self.OnMotion)
    
        self.mToolbar = Toolbar(self.mCanvas)
        self.mToolbar.Realize()
        
        mass_txt = wx.StaticText(self.mToolbar, label='x - y', pos=(230, 7), size=(27, 17))
        mass_txt.SetBackgroundColour("light gray")
        
        self.mass = wx.TextCtrl(self.mToolbar, pos=(260,4), size=(80, 22), style=wx.TE_READONLY)
        
        pageSizer.Add(self.mCanvas, 1, wx.EXPAND)
        pageSizer.Add(self.mToolbar, 0, wx.LEFT | wx.EXPAND)
        
        
        self.mPage.SetSizer(pageSizer)
        self.mPage.Layout()
        self.mParent.AddPage(self.mPage, page_name, False)
        
        pass
    
    def OnMotion(self, evt):
        
        if evt.inaxes:
            xpos = evt.xdata
            ypos = evt.ydata
            self.mass.SetValue('(%d,%d)' % (xpos, ypos))

        pass
    
    def DrawXYplot(self, x_axis, y_axis, y_label):
        #str(y_label[ind])
        ax = self.mFig.add_subplot(1,1,1)
        
        for ind, row in enumerate(y_axis):
            ax.plot(x_axis, row, 'o-', label=str(y_label[ind]))
            pass
       
         
        ax.grid(which='both') 
        ax.legend()
        
        pass
    
    def DrawDotLine(self, x_axis, y_axis):
        _x = np.linspace(int(x_axis[:1][0]), int(x_axis[-1:][0]))
        
        ax = self.mFig.add_subplot(1,1,1)
        for row in y_axis:
            ax.plot(range(len(x_axis)), row, 'yo-')
        
        
        pass

class ParseDataPage(wx.Panel):
    def __init__(self, parent, page_name):
        self.mPage = wx.Panel(parent, 
                         wx.ID_ANY, 
                         wx.DefaultPosition, 
                         wx.DefaultSize, 
                         wx.TAB_TRAVERSAL)
        pageSizer = wx.BoxSizer(wx.VERTICAL)
        
        
        self.mPage.SetSizer(pageSizer)
        self.mPage.Layout()
        parent.AddPage(self.mPage, page_name, False)
        
        pass

class GridDataPage(wx.Panel):
    def __init__(self, parent, page_name, filename, cmd, grid_row_limit, isMp, isParsing, timezone,  args, args_index):
        ##Panel Setting
        self.mParent = parent
        self.mPage = wx.Panel(parent, 
                         wx.ID_ANY, 
                         wx.DefaultPosition, 
                         wx.DefaultSize, 
                         wx.TAB_TRAVERSAL)
        
        pageSizer = wx.BoxSizer(wx.VERTICAL)
        statusSizer = wx.BoxSizer(wx.HORIZONTAL)
        gridSizer = wx.BoxSizer(wx.VERTICAL)
        spinnerSizer = wx.BoxSizer(wx.HORIZONTAL)
        exportSizer = wx.BoxSizer(wx.VERTICAL)
        pageSizer.Add( statusSizer, 0, wx.EXPAND, 5 )
        pageSizer.Add( gridSizer, 0, wx.EXPAND, 5 )
        statusSizer.Add(spinnerSizer, 1, wx.EXPAND, 5)
        statusSizer.Add(exportSizer, 0, 0, 5)
        view.g_progress.SetValue(0)
        '''
        self.m_model = None
        self.mPage.cmd = ''
        self.mPage.condition = ''
        self.mPage.page_name = ''
        self.mPage.filename = ''
        self.mPage.m_data_offset = 0
        self.mPage.m_data_row_limit = 0
        self.mPage.m_data = None
        
        self.mPage.m_data_len = 0
        self.mPage.m_grid_title = ''
        self.mPage.m_grid_title_len = 0
        #self.mPage.m_data = self.mPage.m_data[1]
        
        self.mPage.m_current_page_number = 0
        self.mPage.m_max_page_number = 0
        self.mPage.show_page_number = 0
        self.mPage.m_min_page_number = 1
        '''
        
        self.IsParsing = bool(isParsing)
        #'''
        self.m_model = Model(filename)
        self.m_model.SetIsParsing(self.IsParsing)
        self.m_model.SetTimeZone(timezone)
        self.m_model.SetIsMultiProcessing(isMp)
        self.m_progress = args
        
        
        self.mPage.cmd = cmd
        self.mPage.condition = cmd.partition(' WHERE ')[2]
        self.mPage.page_name = page_name
        #mPage.m_id = 
        #init data preview
        self.mPage.filename = filename
        self.mPage.m_data_offset = 0
        self.mPage.m_data_row_limit = grid_row_limit
        #'''
        self.mPage.timezone = timezone
        
        view.g_progress.SetValue(20)
        self.mPage.m_data = self.m_model.GetRowRangeData(self.mPage.cmd,
                                                         self.mPage.m_data_row_limit, 
                                                         self.mPage.m_data_offset)
        
        
        #'''
        '''
        Thd.Thread(target=self.m_model.GetRowRangeData, args=(self.mPage.cmd,
                                                              self.mPage.m_data_row_limit, 
                                                              self.mPage.m_data_offset)).start()
        #global model.callback_data
        self.mPage.m_data = model.callback_data
        #'''
        
        if not self.mPage.m_data:
            return
        
        self.mPage.m_data_len = self.m_model.parse.find_total_row_number(self.mPage.cmd)
        self.mPage.m_grid_title = self.mPage.m_data[0]
        self.mPage.m_grid_title_len = len(self.mPage.m_data[0])
        #self.mPage.m_data = self.mPage.m_data[1]
        
        self.mPage.m_current_page_number = 1
        
        self.mPage.m_max_page_number = 0
        
        if self.mPage.m_data_len > self.mPage.m_data_row_limit:
            self.mPage.m_max_page_number = self.mPage.m_data_len//self.mPage.m_data_row_limit
            if not self.mPage.m_data_len/self.mPage.m_data_row_limit == 0:
                self.mPage.m_max_page_number += 1
            else:
                pass
        else:
            self.mPage.m_max_page_number = 1
            pass 
        self.mPage.show_page_number = '(  %s  /  %s   )'%(self.mPage.m_current_page_number, self.mPage.m_max_page_number)
        
        self.mPage.m_min_page_number = 1
        #'''
        self.spinText = wx.TextCtrl(self.mPage,
                                    wx.ID_ANY,
                                    wx.EmptyString,
                                    wx.DefaultPosition,
                                    wx.DefaultSize,
                                    wx.TE_READONLY)
        
        
        self.spinText.SetValue(str(self.mPage.show_page_number))
        
        self.spinButton = wx.SpinButton(self.mPage, 
                                        wx.ID_ANY,
                                        wx.DefaultPosition, 
                                        wx.DefaultSize,
                                        wx.SP_WRAP)
        '''
        spinCtrl = wx.SpinCtrl(mPage, 
                                 wx.ID_ANY, wx.EmptyString, 
                                 wx.DefaultPosition, 
                                 wx.DefaultSize, 
                                 wx.SP_WRAP, 
                                 1, mPage.m_data_len, 1)
        '''
        
        staticText = wx.StaticText( self.mPage, 
                                         wx.ID_ANY, 
                                         u"Total Rows: " + str(self.mPage.m_data_len), 
                                         wx.DefaultPosition, 
                                         wx.DefaultSize, 0 )
        vst = wx.StaticLine( self.mPage, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, wx.LI_VERTICAL )
        
        staticText1 = wx.StaticText( self.mPage, 
                                         wx.ID_ANY, 
                                         u"Command: " + str(self.mPage.cmd), 
                                         wx.DefaultPosition, 
                                         wx.DefaultSize, 0 )
        
        
        
        exportButton = wx.Button( self.mPage, 
                                  wx.ID_ANY, 
                                  u"EXPORT", 
                                  wx.DefaultPosition, 
                                  wx.DefaultSize, 0 )
        
        
        spinnerSizer.Add(self.spinText, 0, wx.EXPAND, 5)
        spinnerSizer.Add(self.spinButton, 0, wx.EXPAND, 5)
        spinnerSizer.Add(staticText, 0, wx.ALL, 5)
        spinnerSizer.Add( vst, 0, wx.EXPAND|wx.ALIGN_CENTER_VERTICAL|wx.ALL, 5 )
        spinnerSizer.Add( staticText1, 0, wx.ALL, 5 )
        exportSizer.Add(exportButton, 0, wx.ALIGN_RIGHT, 5)
        
        #row_size = len(data)
        #col_size = len(data[0])
        
        self.mGrid = HugeTableGrid(self.mPage, 
                                   self.mPage.m_data[0], 
                                   self.mPage.m_data,
                                   len(self.mPage.m_data), 
                                   len(self.mPage.m_data[0]))
        #self.mGrid.Refresh()
        
        gridSizer.Add( self.mGrid, 1, wx.EXPAND, 5 )
        
        self.mPage.SetSizer(pageSizer)
        self.mPage.Layout()
        parent.AddPage(self.mPage, page_name, False)     
        
        view.g_progress.SetValue(100)
        View.Info("Process has Done!")
        
        parent.Bind(wx.EVT_SPIN_UP, self.OnSpinUp, self.spinButton)
        parent.Bind(wx.EVT_SPIN_DOWN, self.OnSpinDown, self.spinButton)
        parent.Bind(wx.EVT_BUTTON, self.OnExportExcel, exportButton)
        
        
        self.SelectedAIOColLabel = list()
        self.mGrid.Bind(wx.grid.EVT_GRID_LABEL_RIGHT_CLICK, self.OnShowPopupMenu)
        self.mGrid.Bind(wx.grid.EVT_GRID_RANGE_SELECT, self.OnDragSelection)
        self.mGrid.Bind(wx.grid.EVT_GRID_SELECT_CELL, self.OnSingleSelect)
        #self.mGrid.GetGridWindow().Bind(wx.grid.EVT_GRID_LABEL_RIGHT_CLICK, self.OnShowPopupMenu)
        
        pass
    
    
    
    
    #callback export excel
    
    def OnExportExcel(self, evt):
        _xlax_name = self.mPage.page_name.replace('.db', '.xlsx').replace(' ', '_').replace(':', '_')
        _col_offset = 0
        _row_offset = 0
        _data = self.mPage.m_data
        
        
        
        
        try:
            _doc = openpyxl.Workbook()
            #_doc.save(_xlax_name)
        except:
            View.Warring("Excel Create Error!")
            pass
        
       
        
        #threading.Thread()
        #_doc = openpyxl.load_workbook(_xlax_name)
        _sht_active = _doc.active
        
        
        
        for row_index, row_data in enumerate(_data):
            _sht_active.append(row_data)
        
        
        '''
            for col_index, value_t in enumerate(row_data):
                
                _sht_active.cell(column= _col_offset + col_index + 1, 
                                 row= _row_offset + row_index + 1, 
                                 value = value_t)
        '''
                #_sht_active.append(value_t)
        
        
        try:
            
            _t =threading.Thread(target= lambda doc, xlax_name: doc.save(xlax_name) , args={_doc, _xlax_name,})
            _t.start()
            #_doc.save(_xlax_name)
            
            View.Info("Process has Done!")
        except:
            View.Warring("Excel Write Error!")
            pass
        
        
        
        pass
    
    
    def OnExportBigExcel(self, evt):
        _parse = model.LogParse()
        _parse.set_filename(self.mPage.filename)
        _excel_rows_max_limit = 16384
        _excel_offset = 0
        _col_offset = 0
        _row_offset = 0
        
        _data_total_rows = _parse.find_total_row_number(self.mPage.cmd)
        #print 'Data Total Rows is %d'%_data_total_rows
        #print self.mPage.page_name
        _xlax_name = self.mPage.page_name.replace('.db', '.xlsx').replace(' ', '').replace(':', '_')
        #print _xlax_name
        
        #'''count 
        if _data_total_rows >= _excel_rows_max_limit:
            _sheet_count = _data_total_rows/_excel_rows_max_limit
            if not _data_total_rows % _excel_rows_max_limit == 0:
                _sheet_count+=1
            else:
                pass
        else:
            _sheet_count = 1
            pass
        #'''
        
        _doc = openpyxl.Workbook()
        _doc.save(_xlax_name)
        
        #_data = self.mPage.m_data
        
        '''
        for row_index, row_data in enumerate(_data):
            for col_index, value_t in enumerate(row_data):
                _sht_active.cell(column = _col_offset + col_index + 1,
                                row= _row_offset + row_index + 1,
                                value = value_t)
        #'''
        #print self.mPage.cmd
        #'''
        
        
        for _sht_index in range(_sheet_count):
            
            _doc = openpyxl.load_workbook(_xlax_name, write_only = True)
            #_sht_active = _doc.active
            
            _sht_active = _doc.create_sheet(title=str(_sht_index))
            _sht_active.title = str(_sht_index)
            
            #self.m_model.GetRowRangeData()
            #get data
            _data = self.m_model.GetRowRangeData(self.mPage.cmd, _excel_rows_max_limit, _excel_offset)
            print _data[0]
            
            #write to excel
            for row_index, row_data in enumerate(_data):
                for col_index, value_t in enumerate(row_data):
                    _sht_active.cell(column = _col_offset + col_index + 1,
                                    row= _row_offset + row_index + 1,
                                    value = value_t)
            
            _excel_offset += _excel_rows_max_limit
            _doc.save(_xlax_name)
            print _sht_index
            #_sht_active = _doc.create_sheet()
            #progress_value += progress_value_offset
        
        #'''
        
        
        
        View.Info("Process has Done!")
        pass
    
    def OnExportCSV(self, evt):
        view.g_progress.SetValue(0)
        projectName = self.mPage.page_name.partition('.d')[0] 
        #exportFileName = projectName.replace(': ', '_') + '.xlsx'
        exportFileName = projectName.replace(': ', '_') + '.csv'
        #excel_rows_max_limit = 8192
        #excel_offset = 0
        #sheet_count = 0
        #progress_value = 10
        #progress_value_offset = 0
        #sheet_index = 0
        
        #start_row = 0
        #start_col = 0
        #self.m_progress.SetValue(int(10))
        
        doc = open(exportFileName, 'w')
        
        data_total_rows = self.m_model.parse.find_total_row_number( self.mPage.cmd)
        view.g_progress.SetValue(10)
        
        #self.m_progress.SetValue(int(20))
        
        #self.m_progress.SetValue(int(70))
        
        for row in self.mPage.m_data:
            for value in row:
                doc.write(str(value))
                doc.write(str(','))
            doc.write(str('\n'))
        view.g_progress.SetValue(80)
        #self.m_progress.SetValue(int(90))
        doc.close()
        view.g_progress.SetValue(100)
        View.Info("Process has Done!")
        
        
        '''
        data_total_rows = self.parse.find_total_row_number(self.mPage.filename, 
                                                           self.mPage.cmd)
        
        if data_total_rows >= excel_rows_max_limit:
            sheet_count = data_total_rows/excel_rows_max_limit
            if not data_total_rows % excel_rows_max_limit == 0:
                sheet_count+=1
            else:
                pass
        else:
            sheet_count = 1
            pass
        progress_value_offset = 80/sheet_count
        
        doc = openpyxl.Workbook()
        sht_active = doc.active
        for sht_index in range(sheet_count):
            sht_active.title = str(sht_index)
            
            data = self.m_model.GetRowRangeData(self.mPage.filename,
                                                self.mPage.cmd,
                                                excel_rows_max_limit, 
                                                excel_offset)
            
            for row_index, row_data in enumerate(data):
                
                for col_index, value_t in enumerate(row_data):
                    sht_active.cell(column = start_col + col_index + 1,
                                    row= start_row + row_index + 1,
                                    value = value_t)
            excel_offset += excel_rows_max_limit
            sht_active = doc.create_sheet()
            progress_value += progress_value_offset
            self.m_progress.SetValue(progress_value)
        
        doc.save(exportFileName)
        '''
       
        #self.m_progress.SetValue(int(100))
        
        pass
    
    
    def OnSpinUp(self, evt):
        view.g_progress.SetValue(0)
        row_limit = self.mPage.m_data_row_limit
        if self.mPage.m_current_page_number <= self.mPage.m_max_page_number and self.mPage.m_current_page_number >= self.mPage.m_min_page_number:
            if self.mPage.m_current_page_number == self.mPage.m_max_page_number:
                self.mPage.m_current_page_number = self.mPage.m_min_page_number
            else:
                self.mPage.m_current_page_number += 1
                #self.mPage.m_data_offset += self.mPage.m_data_row_limit
        else:
            pass
        self.mPage.m_data_offset = (self.mPage.m_current_page_number-1)*self.mPage.m_data_row_limit
        
        if self.mPage.m_current_page_number == self.mPage.m_max_page_number:
            row_limit = (self.mPage.m_current_page_number * self.mPage.m_data_row_limit) - self.mPage.m_data_len
            
        else:
            pass
        
        self.mPage.m_data = self.m_model.GetRowRangeData(self.mPage.cmd, 
                                                         row_limit, 
                                                         self.mPage.m_data_offset)
        self.mPage.show_page_number = '(  %s  /  %s   )'%(self.mPage.m_current_page_number, self.mPage.m_max_page_number)
        self.spinText.SetValue(str(self.mPage.show_page_number))
        view.g_progress.SetValue(20)
        self.mGrid.OnUpdate(self.mPage.m_data[0], 
                           self.mPage.m_data, 
                           len(self.mPage.m_data), 
                           len(self.mPage.m_data[0]))
        view.g_progress.SetValue(100)
        View.Info("Process has Done!")
        pass
    
    def OnSpinDown(self, evt):
        view.g_progress.SetValue(0)
        row_limit = self.mPage.m_data_row_limit
        if self.mPage.m_current_page_number <= self.mPage.m_max_page_number and self.mPage.m_current_page_number >= self.mPage.m_min_page_number:
            if self.mPage.m_current_page_number == self.mPage.m_min_page_number:
                self.mPage.m_current_page_number = self.mPage.m_max_page_number
                
            else:
                self.mPage.m_current_page_number -= 1
        else:
            pass
        
        self.mPage.m_data_offset = (self.mPage.m_current_page_number-1)*self.mPage.m_data_row_limit
        
        if self.mPage.m_current_page_number == self.mPage.m_max_page_number:
            row_limit = (self.mPage.m_current_page_number * self.mPage.m_data_row_limit) - self.mPage.m_data_len
            
        else:
            pass
        
        self.mPage.m_data = self.m_model.GetRowRangeData(self.mPage.cmd, 
                                                         row_limit, 
                                                         self.mPage.m_data_offset)
        self.mPage.show_page_number = '(  %s  /  %s   )'%(self.mPage.m_current_page_number, self.mPage.m_max_page_number)
        self.spinText.SetValue(str(self.mPage.show_page_number))
        view.g_progress.SetValue(20)
        self.mGrid.OnUpdate(self.mPage.m_data[0], 
                            self.mPage.m_data,
                            len(self.mPage.m_data), 
                            len(self.mPage.m_data[0]))
        view.g_progress.SetValue(100)
        View.Info("Process has Done!")
        
        
        
        pass
    
    
    def OnSingleSelect(self, evt):
        cells = self.mGrid.GetSelectedCells()
        if not cells:
            if self.mGrid.GetSelectionBlockTopLeft():
                top_left = self.mGrid.GetSelectionBlockTopLeft()[0]
                bottom_right = self.mGrid.GetSelectionBlockBottomRight()[0]
                #print top_left, bottom_right
            else:
                pass
        else:
            #print cells
            pass
        pass
    
    
    def OnDragSelection(self, evt):
        if not evt.Selecting():
            temp = self.SelectedAIOColLabel
            del self.SelectedAIOColLabel[:]
            self.SelectedAIOColLabel = temp
            return
        
        top_left = evt.GetTopLeftCoords()
        bottom_right = evt.GetBottomRightCoords()
        
        
        
        
        if bottom_right[0] == self.mGrid.GetNumberRows()-1:#col label selected
            if evt.Selecting():
                self.SelectedAIOColLabel.append(bottom_right[1])
               
            pass
        
        pass
    
    
    def OnShowPopupMenu(self, evt):
        self.popup_label = evt.GetCol()
        self.popup_menu_title = list()
       
        
        if not hasattr(self, 'popupID1'):
            self.popupID1 = wx.NewId()
            self.popupID2 = wx.NewId()
            self.popupID3 = wx.NewId()
        
        
        
        menu = wx.Menu()
        
        if evt.GetRow() == -1:#right click col label
            if len(self.SelectedAIOColLabel) > 1:#multi select
                pass
            else:
                pass
            
            for ind in range(len(self.SelectedAIOColLabel)):
                label_name = str(self.mGrid.GetColLabelValue(self.SelectedAIOColLabel[ind]))
                if not ('ai' in label_name.lower() or 'ao' in label_name.lower()):
                    return
                self.popup_menu_title.append(label_name)
            
            menu.SetTitle(str(self.popup_menu_title))
            item_plot_1 = wx.MenuItem(menu, self.popupID1, 'Draw X-Y plot', '', wx.ITEM_CHECK)
            item_plot_2 = wx.MenuItem(menu, self.popupID2, 'Draw Dot&Line plot', '', wx.ITEM_CHECK)
            menu.AppendItem(item_plot_1)
            menu.AppendItem(item_plot_2)
            menu.Bind(wx.EVT_MENU, self.OnXYPlot, item_plot_1)
            menu.Bind(wx.EVT_MENU, self.OnDotLine, item_plot_2)
            '''
            if len(self.SelectedAIOColLabel) > 1:#multi select
                
                for ind in range(len(self.SelectedAIOColLabel)):
                    label_name = str(self.mGrid.GetColLabelValue(self.SelectedAIOColLabel[ind]))
                    if not ('ai' in label_name.lower() or 'ao' in label_name.lower()):
                        return
                    self.popup_menu_title.append(label_name)
                
                menu.SetTitle(str(self.popup_menu_title))
                item_plot = wx.MenuItem(menu, self.popupID1, 'Draw X-Y plot', '', wx.ITEM_CHECK)
                menu.AppendItem(item_plot)
                menu.Bind(wx.EVT_MENU, self.OnXYPlot, item_plot)
            
            else:
                label_name = self.mGrid.GetColLabelValue(self.popup_label)
                self.popup_menu_title = label_name
                menu.SetTitle(self.popup_menu_title)
                
                # is col ai or ao
                if 'ai' in str(label_name).lower() or 'ao' in str(label_name).lower():
                    self.SelectedAIOColLabel.append(self.popup_label)
                    item_plot = wx.MenuItem(menu, self.popupID1, 'Draw X-Y plot', '', wx.ITEM_CHECK)
                    menu.AppendItem(item_plot)
                    menu.Bind(wx.EVT_MENU, self.OnXYPlot, item_plot)
                    pass
            '''
            pass
        
        if evt.GetCol() == -1:#right click row label
            print 'row label'
            pass
        
        
        self.mParent.GetParent().PopupMenu(menu)
        menu.Destroy()
       
        pass
    
    
    def OnXYPlot(self, evt):
        #temp = self.popup_menu_title
        #del self.popup_menu_title
        #self.popup_menu_title = temp
        
        
        rows = int(self.mGrid.GetNumberRows())
        row_aio_data = [[] for _ in range(len(self.SelectedAIOColLabel))]
        row_label_data = list()
        
        if len(self.SelectedAIOColLabel) > 0:
            for ind, col in enumerate(self.SelectedAIOColLabel):
                temp = list()
                row_label_data.append(str(self.mGrid.GetColLabelValue(col)))
                for i in range(rows):
                    value = self.mGrid.GetCellValue(i, col)
                    if not value:
                        value = 0
                    temp.append(float(value))
                row_aio_data[ind] = temp
            
            pass
        else:
            pass
        
        '''
        if len(self.SelectedAIOColLabel) > 1:
            for col in self.SelectedAIOColLabel:
                for i in range(rows):
                    value = self.mGrid.GetCellValue(i, col)
                    if not value:
                        value = 0
                    temp.append(float(value))
            row_aio_data.append(temp)
            
            pass
        else:
            
            for i in range(rows):
                value = self.mGrid.GetCellValue(i, self.popup_label)
                if not value:
                    value = 0
                temp.append(float(value))
            
            row_aio_data.append(temp)
        '''
        
        page_name = 'X-Y Plot: %s'%self.popup_menu_title
        
        PlotDataPage(self.mParent, page_name).DrawXYplot(x_axis=range(rows), 
                                                         y_axis=row_aio_data,
                                                         y_label=row_label_data,)
        
        
        
        pass
    
    def OnDotLine(self, evt):
        
        x_axis = list()
        row_aio_data = [[] for _ in range(len(self.SelectedAIOColLabel))]
        rows = int(self.mGrid.GetNumberRows())
        
        #get timestamp as x_axis
        label_timestamp = 0
        if 'timestamp' in str(self.mGrid.GetColLabelValue(label_timestamp)).lower():
            for i in range(rows):
                value = self.mGrid.GetCellValue(i, label_timestamp)
                if not value:
                    value = 0
                x_axis.append(float(value))
            
            pass
        
        if len(self.SelectedAIOColLabel) > 0:
            for ind, col in enumerate(self.SelectedAIOColLabel):
                temp = list()
                for i in range(rows):
                    value = self.mGrid.GetCellValue(i, col)
                    if not value:
                        value = 0
                    temp.append(float(value))
                row_aio_data[ind] = temp
            
            pass
        else:
            pass
    
       
        
        
        
        page_name = 'Dot&Line Plot: %s'%self.popup_menu_title
        
        PlotDataPage(self.mParent, page_name).DrawDotLine(x_axis, row_aio_data)
        pass
    
    
class HugeTableGrid(wx.grid.Grid):
    def __init__(self, parent, title, data, rows, cols):
        wx.grid.Grid.__init__(self, parent, -1)
        
        
        self.table = model.HugeTable(title, data, rows, cols)
        self.SetTable(self.table, True)
        
        self.cols = self.GetNumberCols() 
        self.rows = self.GetNumberRows()
        
        
        self.Bind(wx.EVT_KEY_DOWN, self.OnKey)
        self.Bind(wx.grid.EVT_GRID_CELL_RIGHT_CLICK, self.OnCellRightDown)
        self.Bind(wx.grid.EVT_GRID_LABEL_LEFT_DCLICK, self.OnLabelDClick)
    '''
    def Update(self, *args, **kwargs):
        return wx.grid.Grid.Update(self, *args, **kwargs)
    '''
    
    def Refresh(self):
        print 'refresh test'
    
    def OnUpdate(self, title, data, rows, cols):
        self.table.Clear()
        self.table = model.HugeTable(title, data, rows, cols)
        self.SetTable(self.table, True)
        
        
        #self.table.UpdateData(title, data, rows, cols)     
        '''
        if rows != self.rows:
            if rows < self.rows:
                #cc = 
                self.table.DeleteRows(rows, (self.rows - rows))
                
            elif rows > self.rows:
                self.table.AppendRows(self.rows, rows)
            else:
                pass
        else:
            pass
        ''' 
        
        self.Refresh()
        
    def OnKey(self, event):
        # If Ctrl+C is pressed...
        if event.ControlDown() and event.GetKeyCode() == 67:
            self.copy()
        # If Ctrl+V is pressed...
        if event.ControlDown() and event.GetKeyCode() == 86:
            self.paste('clip')
        # If Ctrl+F is pressed
        if event.ControlDown() and event.GetKeyCode() == 70:
            self.search()
        
        # If Ctrl+Z is pressed...
        if event.ControlDown() and event.GetKeyCode() == 90:
            if self.data4undo[2] != '':
                self.paste('undo')
        # If del is pressed...
        if event.GetKeyCode() == 127:
            # Call delete method
            self.delete()
        # Skip other Key events
        if event.GetKeyCode():
            event.Skip()
            return
        
    def search(self):
        print 'hello'
        print View.Question('hello')
        pass
    
    def copy(self):
        # Number of rows and cols
        #print self.GetSelectionBlockBottomRight()
        #print self.GetGridCursorRow()
        #print self.GetGridCursorCol()
        if self.GetSelectionBlockTopLeft() == []:
            rows = 1
            cols = 1
            iscell = True
        else:
            rows = self.GetSelectionBlockBottomRight()[0][0] - self.GetSelectionBlockTopLeft()[0][0] + 1
            cols = self.GetSelectionBlockBottomRight()[0][1] - self.GetSelectionBlockTopLeft()[0][1] + 1
            iscell = False
        # data variable contain text that must be set in the clipboard
        data = ''
        # For each cell in selected range append the cell value in the data variable
        # Tabs '\t' for cols and '\r' for rows
        for r in range(rows):
            for c in range(cols):
                if iscell:
                    data += str(self.GetCellValue(self.GetGridCursorRow() + r, self.GetGridCursorCol() + c))
                else:
                    data += str(self.GetCellValue(self.GetSelectionBlockTopLeft()[0][0] + r, self.GetSelectionBlockTopLeft()[0][1] + c))
                if c < cols - 1:
                    data += '\t'
            data += '\n'
        # Create text data object
        clipboard = wx.TextDataObject()
        # Set data object value
        clipboard.SetText(data)
        # Put the data in the clipboard
        if wx.TheClipboard.Open():
            wx.TheClipboard.SetData(clipboard)
            wx.TheClipboard.Close()
        else:
            wx.MessageBox("Can't open the clipboard", "Error")

    def paste(self, stage):
        if stage == 'clip':
            clipboard = wx.TextDataObject()
            if wx.TheClipboard.Open():
                wx.TheClipboard.GetData(clipboard)
                wx.TheClipboard.Close()
            else:
                wx.MessageBox("Can't open the clipboard", "Error")
            data = clipboard.GetText()
            if self.GetSelectionBlockTopLeft() == []:
                rowstart = self.GetGridCursorRow()
                colstart = self.GetGridCursorCol()
            else:
                rowstart = self.GetSelectionBlockTopLeft()[0][0]
                colstart = self.GetSelectionBlockTopLeft()[0][1]
        elif stage == 'undo':
            data = self.data4undo[2]
            rowstart = self.data4undo[0]
            colstart = self.data4undo[1]
        else:
            wx.MessageBox("Paste method "+stage+" does not exist", "Error")
        text4undo = ''
        # Convert text in a array of lines
        for y, r in enumerate(data.splitlines()):
            # Convert c in a array of text separated by tab
            for x, c in enumerate(r.split('\t')):
                if y + rowstart < self.NumberRows and x + colstart < self.NumberCols :
                    text4undo += str(self.GetCellValue(rowstart + y, colstart + x)) + '\t'
                    self.SetCellValue(rowstart + y, colstart + x, c)
            text4undo = text4undo[:-1] + '\n'
        if stage == 'clip':
            self.data4undo = [rowstart, colstart, text4undo]
        else:
            self.data4undo = [0, 0, '']

    def delete(self):
        # print "Delete method"
        # Number of rows and cols
        if self.GetSelectionBlockTopLeft() == []:
            rows = 1
            cols = 1
        else:
            rows = self.GetSelectionBlockBottomRight()[0][0] - self.GetSelectionBlockTopLeft()[0][0] + 1
            cols = self.GetSelectionBlockBottomRight()[0][1] - self.GetSelectionBlockTopLeft()[0][1] + 1
        # Clear cells contents
        for r in range(rows):
            for c in range(cols):
                if self.GetSelectionBlockTopLeft() == []:
                    self.SetCellValue(self.GetGridCursorRow() + r, self.GetGridCursorCol() + c, '')
                else:
                    self.SetCellValue(self.GetSelectionBlockTopLeft()[0][0] + r, self.GetSelectionBlockTopLeft()[0][1] + c, '')

    
    def OnCellRightDown(self, e):
        print 'hello'
    
    def OnLabelDClick(self, evt):
        col = evt.GetCol()
        if not col >= 0:
            return
        col_size = self.GetColSize(col)
        #self.AutoSizeColumn(col, True)
        #'''
        if col_size == 80:
            self.AutoSizeColumn(col, True)
        else:   
            self.SetColSize(col, 80)
        #'''
        pass





class Controller:
    def __init__(self, app):
        self.m_view = view.View(None)
        #self.m_model = model.Model()
        
        self.show_current_file = None
        self.current_select_page = None
        self.current_select_project_name = None
        self.current_select_file_name = None
        self.current_select_file_path = None
        self.current_select_file_daytime = None
        
        #self.lock = threading.Lock()
        
        self.cmd = 'SELECT * FROM log_raw'
        #grid parameter
        self.grid_max_colume_number = 0
        self.grid_max_row_number = 8192
        
        #event bind
        #key event
        #self.m_view.m_panel.Bind(wx.EVT_KEY_DOWN, self.OnKeyDown)
        #self.m_view.super().Bind(wx.EVT_KEY_DOWN, self.OnKeyDown)
        
        
        self.IsPageMax = False
        #auimanager EVT_AUINOTEBOOK_TAB_RIGHT_DOWN(winid, fn):
        self.m_view.Bind(wx.lib.agw.aui.EVT_AUINOTEBOOK_PAGE_CHANGED, self.OnANBPageChanged, self.m_view.m_auinotebook)
        self.m_view.Bind(wx.lib.agw.aui.EVT_AUINOTEBOOK_TAB_DCLICK , self.OnANBPageMaxing, self.m_view.m_auinotebook)
        self.m_view.Bind(wx.lib.agw.aui.EVT_AUINOTEBOOK_PAGE_CLOSED , self.OnANBPageClosing, self.m_view.m_auinotebook)
        #self.m_view.Bind(wx.aui.EVT__AUINOTEBOOK_TAB_RIGHT_DOWN, self.OnANBPageMaxing, self.m_view.m_auinotebook)
        #wx.aui.EVT
        #treectrl
        self.m_view.Bind(wx.EVT_TREE_SEL_CHANGED, self.OnSelChanged, self.m_view.m_treectrl)
        self.m_view.Bind(wx.EVT_DIRPICKER_CHANGED, self.OnDirSelected, self.m_view.m_dirpicker)
        self.m_view.Bind(wx.EVT_TREE_ITEM_EXPANDED, self.OnItemExpanded, self.m_view.m_treectrl)
        self.m_view.Bind(wx.EVT_TREE_ITEM_EXPANDING, self.OnItemExpanding, self.m_view.m_treectrl)
        self.m_view.Bind(wx.EVT_TREE_ITEM_ACTIVATED, self.OnItemActivated, self.m_view.m_treectrl)
        self.m_view.Bind(wx.EVT_TREE_ITEM_COLLAPSED, self.OnItemCollapsed, self.m_view.m_treectrl)
        
        
        #notebook_panel_01
        #ctrl panel_01 1f
        self.tablename = ''
        self.src_fifter = ''
        self.state_fifter = ''
        self.level_fifter = ''
        self.proj_name = ''
        self.m_view.Bind(wx.EVT_CHOICE, self.OnProjChoice, self.m_view.m_np1_1f_proj)
        self.m_view.Bind(wx.EVT_CHOICE, self.OnTableChoice, self.m_view.m_np1_1f_table)
        self.m_view.Bind(wx.EVT_TEXT, self.OntcSrcChanged, self.m_view.m_np1_1f_tc_src)
        self.m_view.Bind(wx.EVT_TEXT, self.OntcStateChanged, self.m_view.m_np1_1f_tc_state)
        self.m_view.Bind(wx.EVT_TEXT, self.OnLevelSrcChanged, self.m_view.m_np1_1f_tc_level)
        #ctrl panel_01 2f
        self.m_timezone = 0
        self.m_view.Bind(wx.EVT_CHOICE, self.OnTimeZoneChoice, self.m_view.m_np1_2f_timezone)
        self.m_view.m_np1_checkBox_01.SetValue(False)
        self.m_view.m_np1_st_range.SetLabel('SINGLE')
        self.ts_hl = ''
        self.ts_ll = ''
        self.ts_hl_value = 0
        self.ts_ll_value = 0 
        #self.time_range_fifter = ''
        
        #self.m_view.m_np1_choice_03.Disable()
        #self.m_view.m_np1_choice_04.Disable()
        self.m_view.m_np1_2f_tc_hh.SetEditable(False)
        
        self.m_view.Bind(wx.EVT_CHOICE, self.OnNPLLchoice, self.m_view.m_np1_choice_01)
        self.m_view.Bind(wx.EVT_CHOICE, self.OnNPLLchoice, self.m_view.m_np1_choice_02)
        self.m_view.Bind(wx.EVT_TEXT, self.OntcHHChanged, self.m_view.m_np1_2f_tc_hh)
        self.m_view.Bind(wx.EVT_CHOICE, self.OnNPHLchoice, self.m_view.m_np1_choice_03)
        self.m_view.Bind(wx.EVT_CHOICE, self.OnNPHLchoice, self.m_view.m_np1_choice_04)
        self.m_view.Bind(wx.EVT_TEXT, self.OntcLLChanged, self.m_view.m_np1_2f_tc_ll)
        
        self.m_view.Bind(wx.EVT_CHECKBOX, self.OnNPcheckBox, self.m_view.m_np1_checkBox_01)
        
        self.m_view.Bind(wx.EVT_SLIDER, self.OnSLChanged, self.m_view.m_np1_1f_slider)
        #self.m_view.Bind(wx.EVT_BUTTON, self.OnNPbutton, self.m_view.m_np1_button_01)
        
        
        #ctrl panel_01 3f
        self.m_view.m_np1_3f_tc_cmd.SetValue(self.cmd)
        self.like_fifter = ''
        self.m_view.Bind(wx.EVT_TEXT, self.OntcLIKEChanged, self.m_view.m_np1_3f_tc_like)
        
        self.m_view.Bind(wx.EVT_CHECKBOX, self.OncbCMDOnlyReady, self.m_view.m_np1_3f_cb_cmd)
        self.m_view.Bind(wx.EVT_TEXT, self.OntcCMDChanged, self.m_view.m_np1_3f_tc_cmd)
        
        self.m_view.Bind(wx.EVT_BUTTON, self.OnbtnReset, self.m_view.m_np1_3f_btn_reset)
        self.m_view.Bind(wx.EVT_BUTTON, self.OnbtnQuery, self.m_view.m_np1_3f_btn_query)
        
        #self.m_view.Bind(wx.EVT_TEXT, self.OntcCMDChanged, self.m_view.m_np1_3f_tc_cmd)
        
        
        
        #panel #2
        self.m_view.Bind(wx.EVT_BUTTON, self.OnPanel2Test, self.m_view.btn_nb2_test)
        
        
        
        
        
        
        
        self.m_view.Show()
    
    def OnPanel2Test(self, evt):
        
        ParseDataPage(self.m_view.m_auinotebook,
                      'test')
        
        pass
     
    def OnANBPageChanged(self, evt):
        pageID = evt.GetSelection()
        pageName = self.m_view.m_auinotebook.GetPageText(pageID)
        
        self.current_select_page = pageName
        self.m_view.m_statusBar.SetStatusText(self.current_select_page, 1)
        
        #print pageName, pageID
        pass
    
    
    #evt dir picker
    def OnANBPageMaxing(self, evt):
        
        if self.IsPageMax == True:
            self.m_view.m_dirpicker.Show()
            self.m_view.m_treectrl.Show()
            self.m_view.m_notebook.Show()
            self.m_view.m_statusBar.Show()
            self.m_view.Layout()
            self.IsPageMax = False
        elif self.IsPageMax == False:
            self.m_view.m_dirpicker.Hide()
            self.m_view.m_treectrl.Hide()
            self.m_view.m_notebook.Hide()
            self.m_view.m_statusBar.Hide()
            self.m_view.Layout()
            self.IsPageMax = True
        pass
    def OnANBPageClosing(self, evt):
        
        if self.IsPageMax == True and self.m_view.m_auinotebook.GetPageCount() == 0:
            
            self.m_view.m_dirpicker.Show()
            self.m_view.m_treectrl.Show()
            self.m_view.m_notebook.Show()
            self.m_view.Layout()
            self.IsPageMax = False
            pass
        else:
            pass
        
    
    
    def OnDirSelected(self, evt):
        tree = self.m_view.m_treectrl
        tree.DeleteAllItems()
        root = tree.AddRoot(self.m_view.m_dirpicker.GetPath())
        tree.SetItemImage(root, tree.fldridx, wx.TreeItemIcon_Normal)
        tree.SetItemImage(root, tree.fldropenidx, wx.TreeItemIcon_Expanded)
        tree.SetItemHasChildren(root, True)    
        pass  
    def OnSelChanged(self, evt):#selected item
        itemPath = self.mGetPathOnItem(evt.GetItem())
        itemName = self.m_view.m_treectrl.GetItemText(evt.GetItem())
        
        _sql = model.Sqlite()
        if _sql.is_sqlite_file(itemPath):
            self.current_select_file_path = itemPath
            _model = model.Model(itemPath)
            #self.current_select_project_name = _model.parse.find_current_project()
            self.current_select_file_name = itemName
            self.current_select_file_path = itemPath
            #self.show_current_file = '%s: %s'%(self.current_select_project_name, itemName)
            
            self.show_current_file = '%s'%(itemName)
            self.m_view.m_statusBar.SetStatusText(self.show_current_file, 0)
            #get file daytime
            self.current_select_file_daytime = _model.parse.find_current_project_daytime()
            
            
            #get/update table name
            tablename = ['Table Name']
            tables = _sql.execute_command(itemPath, "SELECT name FROM sqlite_master WHERE type = 'table'")
            for value in tables:
                tablename.append(str(value[0]))
            self.m_view.m_np1_1f_table.SetItems(tablename)
            self.m_view.m_np1_1f_table.SetSelection(0)
        else:
            pass  
    
    def OnItemCollapsed(self, evt):
        item = evt.GetItem()
        self.m_view.m_treectrl.DeleteChildren(item)
        pass    
    def OnItemActivated(self, evt):
        tree = self.m_view.m_treectrl
        itemName = self.m_view.m_treectrl.GetItemText(evt.GetItem())
        itemPath = self.mGetPathOnItem(evt.GetItem())
        
        
        self.current_select_file_path = itemPath
        self.current_select_file_name = itemName
        
        if os.path.isdir(itemPath):
            pass
        else:
            
            
            _sql = model.Sqlite()
            if _sql.is_sqlite_file(itemPath):
                _model = model.Model(itemPath)
                
                #setting model
                
                
                projName = 'preview'
                #projName = 'Preview'
                self.current_select_project_name = projName
                self.show_current_file = '%s: %s'%(projName, itemName)
                page_name = '%s: %s'%(projName, itemName)
                self.m_view.m_statusBar.SetStatusText(self.show_current_file, 0)
                
                isParsing = False#_model.SetIsParsing(0)#mp_message don`t parse
                isMP = False#multiprocess
                GridDataPage(self.m_view.m_auinotebook, 
                             page_name, 
                             itemPath,
                             _model.GetFullDataCmd(),
                             self.grid_max_row_number,
                             isMP,
                             isParsing,
                             self.m_timezone,
                             0,
                             0)
                
                
             
        pass
    def OnItemExpanded(self, evt):
        pass     
    def OnItemExpanding(self, evt):
        item = evt.GetItem()
        tree = self.m_view.m_treectrl
        path = self.mGetPathOnItem(item)
        
        if os.path.isdir(path):
            for dirPath, dirNames, fileNames in os.walk(path):
                for dirName in dirNames:
                    newNode = tree.AppendItem(item, dirName)
                    tree.SetItemImage(newNode, tree.fldridx, wx.TreeItemIcon_Normal)
                    tree.SetItemImage(newNode, tree.fldropenidx, wx.TreeItemIcon_Expanded)
                    tree.SetItemHasChildren(newNode, True)
                
                for fileName in fileNames:
                    newItem = tree.AppendItem(item, fileName)
                    tree.SetItemImage(newItem, tree.fileidx, wx.TreeItemIcon_Normal)
                break
        else:
            pass
        pass
    #evt notebook panel 01
    
        
    def OnProjChoice(self, evt):
        self.proj_name = evt.GetString()
        model.project_name = self.proj_name
        pass
    
    def OnTableChoice(self, evt):
        self.tablename = evt.GetString()
        self.OntcCMDUpdate()
        pass
    
    def OntcSrcChanged(self, evt):
        
        text = evt.GetString()
        self.src_fifter = ''
        if text:
            for value in text.split(','):
                if not value == '' :
                    self.src_fifter += 'src=%s'%value
                    self.src_fifter += ' OR '
                else:
                    pass
            self.src_fifter = '(' + self.src_fifter[:-4] + ')'
        
        self.OntcCMDUpdate()
        
        pass
    def OntcStateChanged(self, evt):
        
        text = evt.GetString()
        self.state_fifter = ''
        if text:
            for value in text.split(','):
                if not value == '' :
                    self.state_fifter += 'state=%s'%value
                    self.state_fifter += ' OR '
                else:
                    pass
            self.state_fifter = '(' + self.state_fifter[:-4] + ')'
        
        self.OntcCMDUpdate()
        pass
    def OnLevelSrcChanged(self, evt):
        
        text = evt.GetString()
        self.level_fifter = ''
        if text:
            for value in text.split(','):
                if not value == '' :
                    self.level_fifter += 'level=%s'%value
                    self.level_fifter += ' OR '
                else:
                    pass
            self.level_fifter = '(' + self.level_fifter[:-4] + ')'
        
        self.OntcCMDUpdate()
        pass
    
    def OntcLIKEChanged(self, evt):
        text = evt.GetString()
        self.like_fifter = ''
        if text:
            for value in text.split(','):
                if not value == '' :
                    self.like_fifter += "msg LIKE '%%%s%%'"%value
                    self.like_fifter += ' OR '
                else:
                    pass
            self.like_fifter = '(' + self.like_fifter[:-4] + ')'
        
        self.OntcCMDUpdate()
        
        pass
    
    def OncbCMDOnlyReady(self, evt):
        #if evt.getValue()
        if evt.IsChecked():
            self.m_view.m_np1_1f_tc_src.SetEditable(False)
            self.m_view.m_np1_1f_tc_state.SetEditable(False)
            self.m_view.m_np1_1f_tc_level.SetEditable(False)
            self.m_view.m_np1_3f_tc_like.SetEditable(False)
            
            self.m_view.m_np1_checkBox_01.Disable()
            self.m_view.m_np1_checkBox_01.SetValue(False)
            self.m_view.m_np1_choice_01.Disable()
            self.m_view.m_np1_choice_02.Disable()
            self.m_view.m_np1_choice_03.Disable()
            self.m_view.m_np1_choice_04.Disable()
            self.m_view.m_np1_2f_tc_ll.SetEditable(False)
            self.m_view.m_np1_2f_tc_hh.SetEditable(False)
            
            self.m_view.m_np1_3f_tc_cmd.SetEditable(True)
        else:
            
            self.m_view.m_np1_1f_tc_src.SetEditable(True)
            self.m_view.m_np1_1f_tc_state.SetEditable(True)
            self.m_view.m_np1_1f_tc_level.SetEditable(True)
            self.m_view.m_np1_3f_tc_like.SetEditable(True)
            
            self.m_view.m_np1_checkBox_01.Enable()
            self.m_view.m_np1_checkBox_01.SetValue(False)
            self.m_view.m_np1_choice_01.Enable()
            self.m_view.m_np1_choice_02.Enable()
            self.m_view.m_np1_choice_03.Enable()
            self.m_view.m_np1_choice_04.Enable()
            self.m_view.m_np1_2f_tc_ll.SetEditable(True)
            self.m_view.m_np1_2f_tc_hh.SetEditable(False)
            
            self.m_view.m_np1_3f_tc_cmd.SetEditable(False)
        
        pass
    def OntcCMDUpdate(self):
        text = self.m_view.m_np1_3f_tc_cmd.GetValue()
        
        tn = ''
        if self.tablename == '' or self.tablename == None or self.tablename == 'Table Name':
            pass
        else:
            tn = self.tablename
        
        src = self.src_fifter
        state = self.state_fifter
        lev = self.level_fifter
        
        like = self.like_fifter
        
        #time_range = self.time_range_fifter      
       
        ll = self.ts_ll
        hl = self.ts_hl
        
        
        
       
        #_list =[src, state, lev, like, time_range] 
        _list =[src, state, lev, like, ll, hl] 
        
        base = text.partition(' WHERE ')[0]
        
        fifter = ''
        for value in _list:
            if not value =='' and not value == None:
                if not fifter =='':
                    fifter += ' AND '
                else:#first
                    fifter = fifter + ' WHERE '
                fifter += value
            else:
                pass
        
        base += fifter
        
        if 'WHERE' in base:
            base = base.partition(' FROM ')[0] + ' FROM ' + tn +' WHERE ' + base.partition(' WHERE ')[2] 
        else:
            base = base.partition(' FROM ')[0] + ' FROM ' + tn
        
        
        
        self.m_view.m_np1_3f_tc_cmd.SetValue(base)
        
        pass
    def OntcCMDChanged(self, evt):
        
        text = evt.GetString()
        self.cmd = text
        pass
    
    def OnbtnReset(self, evt):
        self.mReset()
        
        pass
    def OnbtnQuery(self, evt):
        if not self.current_select_file_path:
            
            pass
        else:
            #project name & tablename not null
            #'''
            if not self.proj_name or self.proj_name == 'Project Name':
                View.Warring('Select a Project')
                return
                
            if not self.tablename or self.tablename == 'Table Name':
                View.Warring('Select a Table')
                return
            #'''
            
            page_name = "%s: %s"%(self.proj_name, 
                                  self.current_select_file_name)
            isParsing = True
            isMP = True
            
            if 'format' in self.tablename:
                isMP = False

            #'''
            GridDataPage(self.m_view.m_auinotebook, 
                             page_name, 
                             self.current_select_file_path, 
                             self.cmd, 
                             self.grid_max_row_number,
                             isMP,
                             isParsing,
                             self.m_timezone,
                             0,
                             0)
            #'''
        pass
        
        
    def OntcLLChanged(self, evt):
        text = evt.GetString()
        
        pass
    def OntcHHChanged(self, evt):
        text = evt.GetString()
        
        pass
        
    def OnTimeZoneChoice(self, evt):
        selected_str = self.m_view.m_np1_2f_timezone.GetStringSelection()[4:-3]
        #print selected_str, int(selected_str)
        self.m_timezone = int(selected_str)
        pass
    
    def OnNPLLchoice(self, evt):
        ll_hr = int(self.m_view.m_np1_choice_01.GetStringSelection().split(' ')[0])
        
        ll_hr -= self.m_timezone#timezone offset
        if ll_hr < 0:
            ll_hr += 24
        
        ll_mm = int(self.m_view.m_np1_choice_02.GetStringSelection().split(' ')[0])
        self.ts_ll =''
        
        if (ll_hr or ll_mm) and self.current_select_file_path:
            _model = model.Model(self.current_select_file_path)
            #_model.SetTimeZone(self.m_timezone)
            if self.m_view.m_np1_checkBox_01.IsChecked():#range mode
                string_split = self.current_select_file_daytime.split('-')
                yy = int(string_split[0])
                MM = int(string_split[1])
                dd = int(string_split[2])
                
                
                
                
                self.ts_ll_value = _model.parse.unixtime_covert(yy, MM, dd, ll_hr, ll_mm, 00)
                
                self.ts_ll = '(ts>=' + str(self.ts_ll_value) + '000000' + ')'            
                pass
            else:#SINGLE
                string_split = self.current_select_file_daytime.split('-')
                yy = int(string_split[0])
                MM = int(string_split[1])
                dd = int(string_split[2])
               
                self.ts_ll_value = _model.parse.unixtime_covert(yy, MM, dd, ll_hr, ll_mm, 00)
                
                ll = int(self.ts_ll_value) - int(self.ts_hl_value)
                hh = int(self.ts_ll_value) + int(self.ts_hl_value)
                
                self.ts_ll = '(ts>=' + str(ll) + '000000' + ')' 
                self.ts_hl = '(ts<=' + str(hh) + '000000' + ')' 
                
                pass
        else:
            self.ts_ll =''
            
        
            
        self.OntcCMDUpdate()
        
        
        
        
        
        """
        if (ll_hr or ll_mm) and self.current_select_file_path and self.m_view.m_np1_checkBox_01.IsChecked():
            string_split = self.current_select_file_daytime.split('-')
            yy = int(string_split[0])
            MM = int(string_split[1])
            dd = int(string_split[2])
            
            parse = model.LogParse(self.current_select_file_path)
            self.ts_ll = parse.unixtime_covert(yy, MM, dd, ll_hr, ll_mm, 00)
            
            if (not self.ts_hl=='') and (not self.ts_ll==''):
            #int(self.ts_hl)>int(self.ts_ll)''' :
                
                self.time_range_fifter = self.m_model.GetTimeRangeCmd(self.current_select_file_path, self.ts_hl, self.ts_ll)
                self.time_range_fifter = '(' + self.time_range_fifter.partition(' WHERE ')[2].partition(' ORDER ')[0] + ')' + ' ORDER ' + self.time_range_fifter.partition(' WHERE ')[2].partition(' ORDER ')[2]
                self.OntcCMDUpdate()
            else:
                pass
        else:
            pass
        """
        pass
    def OnNPHLchoice(self, evt):
        hl_hr = int(self.m_view.m_np1_choice_03.GetStringSelection().split(' ')[0])
        hl_mm = int(self.m_view.m_np1_choice_04.GetStringSelection().split(' ')[0])
        self.ts_hl =''
        
        
        if (hl_hr or hl_mm) and self.current_select_file_path:
            if self.m_view.m_np1_checkBox_01.IsChecked():#RANGE
                string_split = self.current_select_file_daytime.split('-')
                yy = int(string_split[0])
                MM = int(string_split[1])
                dd = int(string_split[2])
                
                _model = model.Model(self.current_select_file_path)
                #_model.SetTimeZone(self.m_timezone)
                hl_hr -= self.m_timezone#timezone offset
                if hl_hr < 0:
                    hl_hr += 24
                
                self.ts_hl_value = _model.parse.unixtime_covert(yy, MM, dd, hl_hr, hl_mm, 00)
                self.ts_hl = '(ts<=' + str(self.ts_hl_value) + '000000' + ')'            
                pass
            else:#SINGLE
                self.ts_hl_value = str((hl_hr*60*60) + (hl_mm*60))
                ll = int(self.ts_ll_value) - int(self.ts_hl_value)
                hh = int(self.ts_ll_value) + int(self.ts_hl_value)
                
                self.ts_ll = '(ts>=' + str(ll) + '000000' + ')' 
                self.ts_hl = '(ts<=' + str(hh) + '000000' + ')' 
                pass
        else:
            self.ts_hl =''
            
        
            
        self.OntcCMDUpdate()
        
        
        
       
        """
        if (hl_hr or hl_mm) and self.current_select_file_path and self.m_view.m_np1_checkBox_01.IsChecked():
            string_split = self.current_select_file_daytime.split('-')
            yy = int(string_split[0])
            MM = int(string_split[1])
            dd = int(string_split[2])
            
            parse = model.LogParse(self.current_select_file_path)
            self.ts_hl = parse.unixtime_covert(yy, MM, dd, hl_hr, hl_mm, 00)
            
            if (not self.ts_hl=='') and (not self.ts_ll==''):
                self.time_range_fifter = self.m_model.GetTimeRangeCmd(self.current_select_file_path, self.ts_hl, self.ts_ll)
                self.time_range_fifter = '(' + self.time_range_fifter.partition(' WHERE ')[2].partition(' ORDER ')[0] + ')' + ' ORDER ' + self.time_range_fifter.partition(' WHERE ')[2].partition(' ORDER ')[2]
                self.OntcCMDUpdate()
            else:
                pass
        else:
            pass
        """
        pass
    
    def OnNPbutton(self, evt):
        if not self.current_select_file_path:
            pass
        else:
            string_split = self.current_select_file_daytime.split('-')
            yy = int(string_split[0])
            MM = int(string_split[1])
            dd = int(string_split[2])
            
            parse = model.LogParse(self.current_select_file_path)
            ts_hl = parse.unixtime_covert(yy, MM, dd, self.hl_hr, self.hl_mm, 00)
            ts_ll = parse.unixtime_covert(yy, MM, dd, self.ll_hr, self.ll_mm, 00)
        
            if ts_hl >= ts_ll:
                #data = self.m_model.GetTimeRangeData(self.current_select_file_path, ts_hl, ts_ll)
                cmd = self.m_model.GetTimeRangeCmd(self.current_select_file_path, ts_hl, ts_ll)
                page_name = "%s: %s [%s:%s ~ %s:%s]"%(self.current_select_project_name, 
                                                      self.current_select_file_name,
                                                      self.ll_hr, self.ll_mm,
                                                      self.hl_hr, self.hl_mm)
                '''
                if  not data and len(data) > 1:
                    pass
                else:
                    page_name = "%s: %s [%s:%s ~ %s:%s]"%(self.current_select_project_name, 
                                                          self.current_select_file_name,
                                                          self.ll_hr, self.ll_mm,
                                                          self.hl_hr, self.hl_mm)
                
                    self.mAddGridPage(page_name, data[0], data)
                '''
                GridDataPage(self.m_view.m_auinotebook, 
                             page_name, 
                             self.current_select_file_path, 
                             cmd, 
                             self.grid_max_row_number,
                             self.m_view.m_progress,
                             0)

            else:
                pass
        pass
    def OnSLChanged(self, evt):
        self.grid_max_row_number = evt.GetInt()
        pass
    def OnNPcheckBox(self, evt):
        if  evt.IsChecked():
            
            #self.m_view.m_np1_choice_03.Enable()
            #self.m_view.m_np1_choice_04.Enable()
            self.m_view.m_np1_st_range.SetLabel('RANGE')
            self.m_view.m_np1_st_offset.SetLabel('TO')
            self.m_view.m_np1_2f_tc_hh.SetEditable(True)
            pass
        else:
            
            #self.m_view.m_np1_choice_03.Disable()
            #self.m_view.m_np1_choice_04.Disable()
            self.m_view.m_np1_st_range.SetLabel('SINGLE')
            self.m_view.m_np1_st_offset.SetLabel('OFFSET')
            self.m_view.m_np1_2f_tc_hh.SetEditable(False)
            pass
        pass
        
    
    #@staticmethod
    
    def mReset(self):
        self.cmd = 'SELECT * FROM log_raw'
        
        self.src_fifter = ''
        self.state_fifter = ''
        self.level_fifter = ''
        self.m_view.m_np1_1f_tc_src.SetValue('')
        self.m_view.m_np1_1f_tc_state.SetValue('')
        self.m_view.m_np1_1f_tc_level.SetValue('')
        self.m_view.m_np1_1f_tc_src.SetEditable(True)
        self.m_view.m_np1_1f_tc_state.SetEditable(True)
        self.m_view.m_np1_1f_tc_level.SetEditable(True)
        
        self.m_view.m_np1_2f_timezone.SetSelection( 11 )
        self.m_view.m_np1_choice_01.SetSelection(0)
        self.m_view.m_np1_choice_02.SetSelection(0)
        self.m_view.m_np1_choice_03.SetSelection(0)
        self.m_view.m_np1_choice_04.SetSelection(0)
        self.ts_hl = ''
        self.ts_ll = ''
        self.ts_ll_value = 0
        self.ts_hl_value = 0
        self.time_range_fifter = ''
        self.m_view.m_np1_st_range.SetLabel('SINGLE')
        self.m_view.m_np1_st_offset.SetLabel('OFFSET')
        self.m_view.m_np1_checkBox_01.Enable()
        self.m_view.m_np1_checkBox_01.SetValue(False)
        self.m_view.m_np1_choice_01.Enable()
        self.m_view.m_np1_choice_02.Enable()
        self.m_view.m_np1_choice_03.Enable()
        self.m_view.m_np1_choice_04.Enable()
        self.m_view.m_np1_2f_tc_ll.SetEditable(True)
        self.m_view.m_np1_2f_tc_hh.SetEditable(False)
        
        self.m_view.m_np1_1f_slider.SetValue(8192)
        #self.m_view.m_np1_2f_slider.SetTickFreq(1,0)
        
        self.like_fifter = ''
        self.m_view.m_np1_3f_tc_like.SetValue('')
        self.m_view.m_np1_3f_tc_like.SetEditable(True)
        self.m_view.m_np1_3f_cb_cmd.SetValue(False)
        
        self.m_view.m_np1_3f_tc_cmd.SetValue('SELECT * FROM log_raw')
        
        self.m_view.m_np1_3f_tc_cmd.SetEditable(False)
        
        pass
    
    def mGetPathOnItem(self, item):
        tree = self.m_view.m_treectrl
        currentPath = tree.GetItemText(item)
        traceItem = item
        if item == tree.GetRootItem():
            pass
        else:
            while traceItem != tree.GetRootItem():
                currentPath = os.path.join(tree.GetItemText(tree.GetItemParent(traceItem)), currentPath)
                traceItem = tree.GetItemParent(traceItem)
        return currentPath
                
    
    
    
    
    def OnKeyDown(self, e):
        key = e.GetKeyCode()
        
        if key == wx.WXK_F12:
        
            print 444
            pass 
        
        pass
   
    
    
    
    
























import multiprocessing

if __name__ == '__main__':
    multiprocessing.freeze_support()
    app = wx.App()
    ctrl = Controller(None)
    app.MainLoop()
    pass